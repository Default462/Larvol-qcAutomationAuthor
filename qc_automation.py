#!/usr/bin/env python

import pandas as pd
import re
from Config import *
from datetime import datetime
import datetime
import os
import sys
df = pd.read_excel(source_file)
import numpy as np
df = df.replace(np.nan, '')

source_file = source_file.split('.')[0]

if os.path.exists(f'{source_file}.txt'):
    os.remove(f'{source_file}.txt')
    print(f'deleting priviouslg generated comments {source_file}.txt')

to_write = ''

if not df.columns.tolist() == ['is_paid','source_id', 'manual_id', 'article_title', 'url', 'authors', 'author_affiliation', 'abstract_text', 'date', 'start_time', 'end_time', 'location', 'session_id', 'news_type',
'session_title', 'session_type', 'category', 'sub_category', 'disclosure']:
    print('hi there')
    to_write += '#########################################################################\n\n row header is wrong \n\n#########################################################################'
    sys.exit("row header is wrong")



df = df.astype(str)
is_paid = df['is_paid'].tolist()
source_id = df['source_id'].tolist()
manual_id = df['manual_id'].tolist()
url = df['url'].tolist()
article_title = df['article_title'].tolist()
authors = df['authors'].tolist()
author_affiliation = df['author_affiliation'].tolist()
abstract_text = df['abstract_text'].tolist()
date = df['date'].tolist()
start_time = df['start_time'].tolist()
end_time = df['end_time'].tolist()
location = df['location'].tolist()
session_title = df['session_title'].tolist()
session_type = df['session_type'].tolist()
category = df['category'].tolist()
sub_category = df['sub_category'].tolist()
disclosure = df['disclosure'].tolist()
session_id = df['session_id'].tolist()
news_type = df['news_type'].tolist()


res = {'source_id':source_id,
'manual_id': manual_id,
'url':url,
'article_title':article_title,
'authors':authors,
'author_affiliation':author_affiliation,
'abstract_text':abstract_text,
'date':date,
'start_time':start_time,
'end_time':end_time,
'location':location,
'session_title':session_title,
'session_id':session_id,
'news_type':news_type,
'session_type':session_type,
'category':category,
'sub_category':sub_category,
'disclosure':disclosure}

paitron_msg = ''

    
def strip_str(res):
    msg = '##############################   extra space at start and end    ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            temp_data = data.strip()
            
            if not data == temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'

temp_msg = strip_str(res)
paitron_msg += temp_msg
to_write += temp_msg


def Encoding_str(res):
    msg = '##############################   Encoding string    ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            temp_data = data
            
            if '_x00' in temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'

temp_msg = Encoding_str(res)
paitron_msg += temp_msg
to_write += temp_msg



def mid_format(res):
    msg = '##############################   invalid mid format   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            try:
                temp_data = re.search('(([A-Za-z]+_)+\d+)',data).group(1)
                
                if not data == temp_data:
                    msg += f'{key} ============>  Bad format   {temp_data} ================>Row no. {ct}\n'
            except:
                msg += f'{key} ============>  Bad format   {data} ================>Row no. {ct}\n'

            
    return f'{msg}\n\n'

temp_msg = mid_format({'manual_id': manual_id})
paitron_msg += temp_msg
to_write += temp_msg



def line_breakes(res):
    msg = '##############################   line_breakes    ##################################\n\n'
    
    for key in res:
        ct = 1
        for data in res[key]:
            ct += 1
            
            temp_data = re.sub('\r','',data,flags=re.S)
            temp_data = re.sub('\n','',data,flags=re.S)
            
            if not data == temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'

temp_msg = line_breakes(res)
paitron_msg += temp_msg
to_write += temp_msg


def author_aff_delimiter(res):
    msg = '##############################   Invalid author_aff_delimiter   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            temp_data = re.sub('\s*;\s*','; ',data,flags=re.S)
            
            if not data == temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'  

temp_msg = (author_aff_delimiter({'authors':authors, 'author_affiliation':author_affiliation}))
paitron_msg += temp_msg
to_write += temp_msg


def space_(res):
    msg = '##############################   space in mid time url   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if ' 'in data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'   

temp_msg = space_({'manual_id':manual_id,'url':url, 'start_time':start_time, 'end_time':end_time})
paitron_msg += temp_msg
to_write += temp_msg



def date_format(res):
    msg = '##############################   Invalid date_format   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if not data:
                continue
            try:
                datetime.datetime.strptime(data, '%B %d, %Y')
                
            except:
                msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
            
            else:
                # October 18, 2022
                if not re.search('([A-Z][a-z]+ \d\d, \d\d\d\d)',data):
                    msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'

            
    return f'{msg}\n\n'      

temp_msg = (date_format({'date':date}))
paitron_msg += temp_msg
to_write += temp_msg


def start_end_time_format(res):
    msg = '##############################   Invalid start_end_time_format   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            
            if not data:
                continue
            try:
                datetime.datetime.strptime(data, '%H:%M')
                
            except:
                msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
                
            else:
                if not re.search('(\d\d:\d\d)',data):
                    msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
         
    return f'{msg}\n\n'      

temp_msg = (start_end_time_format({'start_time':start_time,'end_time':end_time}))
paitron_msg += temp_msg
to_write += temp_msg


def start_end_time_les_0_6(res):
    msg = '##############################   time ( 7 =< time >= 0 )   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if not data:
                continue
            try:
                temp_data = data.replace(' ','').replace(':','').strip()
                if int(temp_data) <= 700:
                    msg += f'{key} ============>    this early time not possible {data}   ===========>   Row no. {ct}\n'
            except:
                continue
                        
    return f'{msg}\n\n'      

temp_msg = (start_end_time_les_0_6({'start_time':start_time,'end_time':end_time}))
paitron_msg += temp_msg
to_write += temp_msg

def is_paid_wrong_data(res):
    msg = '############################   Wrong data in is_paid   ##############################\n\n'

    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if data=='No' or data=="Yes":
                continue

            else:
                msg += f'{key} ============>    Wrong data in is_paid {data}   ===========>   Row no. {ct}\n'

        return f'{msg}\n\n'
    

temp_msg = is_paid_wrong_data({'is_pad':is_paid})
paitron_msg += temp_msg
to_write += temp_msg

def start_end_time_morethan_2330(res):
    msg = '##############################   time ( time >= 23:30 )   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if not data:
                continue
            try:

                temp_data = data.replace(' ','').replace(':','').strip()
                if int(temp_data) >= 2330:
                    msg += f'{key} ============>    this late time not possible {data}   ===========>   Row no. {ct}\n'
            except:
                continue
                        
    return f'{msg}\n\n'      

temp_msg = (start_end_time_morethan_2330({'start_time':start_time,'end_time':end_time}))
paitron_msg += temp_msg
to_write += temp_msg


def if_end_time_small(res):
    msg = '##############################   Start time is grater then End time  ##################################\n\n'

    ct = 1
    for x in range(len(res['end_time'])):
        ct += 1
        if res['end_time'][x] == '':
            continue
        
        temp_end = int(''.join(re.findall('\d',res['end_time'][x],flags=re.S)))
        temp_start = int(''.join(re.findall('\d',res['start_time'][x],flags=re.S)))

        if temp_start > temp_end :
                msg += f'end_time ============>   start_time end_time is same    ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n' 

temp_msg = (if_end_time_small({'start_time':start_time,'end_time':end_time}))
paitron_msg += temp_msg
to_write += temp_msg


def invalid_end_time(res):
    msg = '##############################   start_time is blank but end_time is there   ##################################\n\n'

    ct = 1
    for x in range(len(res['start_time'])):
        ct += 1
        if not res['start_time'][x]:
            if res['end_time'][x]:
                msg += f'end_time ============>   end time not possible    ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'      

temp_msg = (invalid_end_time({'start_time':start_time,'end_time':end_time}))
paitron_msg += temp_msg
to_write += temp_msg


def start_time_end_time(res):
    msg = '##############################   start_time end_time is same   ##################################\n\n'

    ct = 1
    for x in range(len(res['start_time'])):
        ct += 1
        if res['start_time'][x]== '':
            continue

        if res['start_time'][x] == res['end_time'][x]:
                msg += f'end_time ============>   start_time end_time is same    ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'      

temp_msg = (start_time_end_time({'start_time':start_time,'end_time':end_time}))
paitron_msg += temp_msg
to_write += temp_msg


def session_id_format(res):
    msg = '##############################   invalid session id format   ##################################\n\n'

    ct = 1
    for x in range(len(res['session_id'])):
        ct += 1
        if res['session_id'][x]== '':
            continue
        
        if not re.search('^S\d+$',res['session_id'][x],flags=re.S):
                msg += f'session_id ============>   invalid session id format {res["session_id"][x]}  ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'      

temp_msg = (session_id_format({'session_id':session_id}))
paitron_msg += temp_msg
to_write += temp_msg

def news_type_format(res):
    msg = '##############################   invalid news_type format   ##################################\n\n'

    ct = 1
    for x in range(len(res['news_type'])):
        ct += 1
        if res['news_type'][x]== '':
            continue
        
        if not re.search('^Session$|^Abstract$',res['news_type'][x],flags=re.S):
                msg += f'news_type ============>   invalid news_type format {res["news_type"][x]}  ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'      

temp_msg = (news_type_format({'news_type':news_type}))
paitron_msg += temp_msg
to_write += temp_msg


def Formula_in_excel(res):
    msg = '##############################   Formula in excel (AS General)  ##################################\n\n'

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    df = pd.read_excel(res)

    workbook = load_workbook(res)
    worksheet = workbook.active

    for column in df.columns:
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        ct = 0
        for cell in worksheet[column_letter]:
            ct += 1
            if cell.data_type == 'f':
                msg += f"Formula found in column '{column}': {cell.value} row number {ct}\n"
        ct = 0
                        
    return f'{msg}\n\n'      

temp_msg = Formula_in_excel(source_file + '.xlsx')
paitron_msg += temp_msg
to_write += temp_msg


def strip_str(res):
    msg = '##############################   Formula in excel (AS Text)    ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            
            temp_data = re.search('(^=[A-Z]+\()',data,flags=re.S)
            if temp_data:
                temp_data = re.search('(^=[A-Z]+\(.*)',data,flags=re.S)
                msg += f'Formula found in column \'{key}\': {temp_data.group(1)} row number {ct}\n'
            
            
    return f'{msg}\n\n'

temp_msg = strip_str(res)
paitron_msg += temp_msg
to_write += temp_msg


def Two_or_more_session_as_new_in_sid(df):
    msg = '##############################   Two or more Session(news_type) for single SID   ##################################\n\n'

    grouped_bk_group = df.groupby('session_id')

    for _, group_df in grouped_bk_group:
        if group_df['news_type'].to_list().count('Session')>1:

            msg += f' For SID \"{_}\" there is two or more Session in news_type coln for single SID please check  \n\n'

    return f'{msg}\n\n'

temp_msg = Two_or_more_session_as_new_in_sid(df)
paitron_msg += temp_msg
to_write += temp_msg


def multiple_session_title_to_single_sid(df):
    msg = '##############################   multiple session_title to single SID   ##################################\n\n'

    grouped_bk_group = df.groupby('session_id')

    for _, group_df in grouped_bk_group:
        if len(set(group_df['session_title'].to_list()))>1:

            msg += f' SID \"{_}\" is pointing to multiple different session titles please check  \n\n'

    return f'{msg}\n\n'


temp_msg = multiple_session_title_to_single_sid(df)
paitron_msg += temp_msg
to_write += temp_msg



# def Single_session_title_multiple_sid(df):
#     msg = '##############################   Single session title multiple sid   ##################################\n\n'

#     grouped_bk_group = df.groupby('session_title')

#     for _, group_df in grouped_bk_group:

#         if group_df['news_type'].to_list().count('Session')==1:

#             if len(set(group_df['session_id'].to_list()))>1:

#                 msg += f' Title \"{_}\" is pointing to multiple different SID please check  \n\n'

#     return f'{msg}\n\n'


# temp_msg = Single_session_title_multiple_sid(df)
# paitron_msg += temp_msg
# to_write += temp_msg


if '' in manual_id:
    temp_msg = '=================>  vacant cell found in manual_id column  <=================\n\n'
    to_write += temp_msg
    
    
if '' in url:
    temp_msg = '=================>  vacant cell found in url column   <=================\n\n'
    to_write += temp_msg
    
if '' in article_title:
    temp_msg = '=================>  vacant cell found in article_title column   <=================\n\n'
    to_write += temp_msg

if '' in is_paid:
    temp_msg = '=================>  vacant cell found in is_paid column   <=================\n\n'
    to_write += temp_msg

manual_id_to_b_unique = df["source_id"].tolist()
source_id_to_b_unique = df["manual_id"].tolist()
article_title_to_b_unique = df["article_title"].tolist()

manual_id_unique_dic = {}
source_id_unique_dic = {}
article_title_unique_dic = {}

for x in range(len(manual_id_to_b_unique)):
    if not manual_id_unique_dic.get(manual_id_to_b_unique[x],''):
        manual_id_unique_dic[manual_id_to_b_unique[x]] = 1
    else: 
        manual_id_unique_dic[manual_id_to_b_unique[x]] += 1

    if not source_id_unique_dic.get(source_id_to_b_unique[x],''):
        source_id_unique_dic[source_id_to_b_unique[x]] = 1
    else: 
        source_id_unique_dic[source_id_to_b_unique[x]] += 1
        
    if not article_title_unique_dic.get(article_title_to_b_unique[x],''):
        article_title_unique_dic[article_title_to_b_unique[x]] = 1
    else: 
        article_title_unique_dic[article_title_to_b_unique[x]] += 1
        
        
to_write_unique_manual_id = ''
to_write_unique_source_id = ''
to_write_unique_article_title = ''


for x,y in manual_id_unique_dic.items():
    if y>1 and x!='':
        to_write_unique_manual_id = to_write_unique_manual_id + f'{x} ===========> total count is {y}\n'
        
        
for x,y in source_id_unique_dic.items():
    if y>1 and x!='':
        to_write_unique_source_id = to_write_unique_source_id + f'{x} ===========> total count is {y}\n'
        
for x,y in article_title_unique_dic.items():
    if y>1 and x!='':
        to_write_unique_article_title = to_write_unique_article_title + f'{x} ===========> total count is {y}\n'

to_write_2 = f'''##############################   Duplicate mid  ##################################\n\n
          mannual_id which are not unique \n{to_write_unique_source_id}\n\n\n\n\n\n
-------------------------------------------------------------------------------------------------------------------------------------------------------------------
#################################################################        Duplicates          ######################################################################
-------------------------------------------------------------------------------------------------------------------------------------------------------------------
          source_id which are not unique \n{to_write_unique_manual_id}\n\n\n\n\n\n
          
          artical_id which are not unique \n{to_write_unique_article_title}'''

to_write_2 = re.sub('1qaz2wsx.*?total count is \d+','',to_write_2)

to_write += '\n' + to_write_2


with open(f"{source_file}_QC_comments.txt",'w',encoding='utf-8') as f:
    f.write(to_write)



if paitron_msg.count('===========>')<1 and len(to_write_unique_source_id)<1:
    print(to_write_unique_source_id)

    os.system("python separate_authors.py")
    print('running separate_authors')
    
    

sponsor = ["Sponsor",
"Sponsored by",
"Sponsorship",
"Fund",
"Funded by",
"Financed",
"Financed by",
"Financial support",
"Supported by",
"Acknowledgement",
"Acknowledged by",
"Registration ID",
"Clinical Trial ID"]


sponsor_msg = ''


def sponsor_col(res):
    msg = '##############################   sponsor   ##################################\n\n'
    
    for key in res:
        ct = 1
        for data in res[key]:
            ct += 1
            
            for sp_key in sponsor:
                if sp_key.lower() in data.lower():

                    msg += f'May can be sponsor {sp_key} ==========> {key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'

temp_msg = sponsor_col({"abstract_text":abstract_text})
sponsor_msg += temp_msg

with open(f'{source_file}_sponsor.txt','w',encoding='utf-8')as s_dtails:
    s_dtails.write(sponsor_msg)


